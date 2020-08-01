from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from calendar import monthrange
from random import randint

#Code that creates a simple work schendule in an spreadsheet document

filename = "Escala de Serviço.xlsx"

wb = load_workbook(filename)
ws =  wb.active

day, month_total_days = monthrange(2020, 8)

guards = {
    1 : {'nome': 'Júnior', 'work_amount': 9},
    2 : {'nome': 'Diolindo', 'work_amount': 9},
    3 : {'nome': 'Solon', 'work_amount': 9},
    4 : {'nome': 'Clênio', 'work_amount': 9},
    5 : {'nome': 'Antônio', 'work_amount': 8},
    6 : {'nome': 'Dorgival', 'work_amount': 9},
    7 : {'nome': 'Fabiano', 'work_amount': 9},
}

for x in range(2, (month_total_days + 2)):
    mod = None

    if (ws[x][0].value % 2 == 0):
        mod = "Par"
    else:
        mod = "Impar"

    for y in range(1, 5):
        if y == 2 or y == 4:
            pass
        else:
            if ws[x][y].value == None:
                selected_driver = None
                while selected_driver == None:
                    number = randint(1,7)
                    aux_driver = guards[number]
                    if ws[x-1][1].value != aux_driver['nome'] and ws[x-1][3].value != aux_driver['nome']:
                        if y == 3:
                            if ws[x][1].value != aux_driver['nome']:
                                if aux_driver['work_amount'] > 0:
                                    if aux_driver['nome'] == 'Júnior':
                                        if mod == 'Impar':
                                            guards[number]['work_amount'] = aux_driver['work_amount'] - 1
                                            selected_driver = aux_driver
                                        else:
                                            pass
                                    else:
                                        guards[number]['work_amount'] = aux_driver['work_amount'] - 1
                                        selected_driver = aux_driver
                                else:
                                    pass
                            else:
                                pass
                        else:
                            if aux_driver['work_amount'] > 0:
                                if aux_driver['nome'] == 'Júnior':
                                    if mod == 'Par':
                                        guards[number]['work_amount'] = aux_driver['work_amount'] - 1
                                        selected_driver = aux_driver
                                    else:
                                        pass
                                else: 
                                    guards[number]['work_amount'] = aux_driver['work_amount'] - 1
                                    selected_driver = aux_driver
                            else:
                                pass

                ws[x][y].value = selected_driver['nome']
                wb.save(filename)
            else: 
                for z in guards:
                    if guards[z]['nome'] == ws[x][y].value:
                        guards[z]['work_amount'] = guards[z]['work_amount'] - 1
        

wb.save(filename)






