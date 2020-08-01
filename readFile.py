from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from calendar import monthrange


#Code created for debug reasons


filename = "Escala de Serviço.xlsx"

wb = load_workbook(filename)
ws =  wb.active

day, month_total_days = monthrange(2020, 8)


guards = {
    1 : {'nome': 'Júnior', 'work_amount': 0},
    2 : {'nome': 'Diolindo', 'work_amount': 0},
    3 : {'nome': 'Solon', 'work_amount': 0},
    4 : {'nome': 'Clênio', 'work_amount': 0},
    5 : {'nome': 'Antônio', 'work_amount': 0},
    6 : {'nome': 'Dorgival', 'work_amount': 0},
    7 : {'nome': 'Fabiano', 'work_amount': 0},
}

for x in range(2, (month_total_days + 2)):
    for y in range(1, 5):
        if y == 2 or y == 4:
            pass
        else:
            if ws[x][y].value != None:
                for z in guards:
                    if guards[z]['nome'] == ws[x][y].value:
                        guards[z]['work_amount'] = guards[z]['work_amount'] + 1

total = 0

for guard in guards:
    total += guards[guard]['work_amount']
    print(guards[guard]['nome'], ': ', guards[guard]['work_amount']) 

print(total)